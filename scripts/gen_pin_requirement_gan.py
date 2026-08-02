"""Derive the experimental full-GAN chip's pin sheet from the main chip's.

Answers "do the pin requirements change now that the discriminator is on chip?".
Short version, and the reason this script exists rather than a hand-edited copy: the
discriminator changes exactly ONE pin.  Everything else that moved is the parallel
burst bus, which is a throughput feature and would have been added to the main chip
too.  Generating the sheet from the original keeps that attribution auditable.

Reads  GAN_CHIP_Pin_Requirement.xlsx           (the main chip, as submitted)
Writes GAN_CHIP_Pin_Requirement_gan.xlsx       (the experimental chip)

The source file is never modified: it is the submitted collateral and has a
Google-Docs twin that has to be re-uploaded by hand.

    python3 scripts/gen_pin_requirement_gan.py
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "GAN_CHIP_Pin_Requirement.xlsx"
DST = ROOT / "GAN_CHIP_Pin_Requirement_gan.xlsx"

# Row numbers in the source sheet (1-based, including the header row).
ROW_MOSI, ROW_MISO = 13, 15
ROW_BUSY, ROW_DONE, ROW_WBDONE = 16, 17, 18
ROW_BIDIR7 = 20                       # bidir[7]; bidir[n] is at ROW_BIDIR7 + (n - 7)
ROW_SUMMARY, ROW_PROTO, ROW_PLACE = 94, 95, 96


def bidir_row(n: int) -> int:
    return ROW_BIDIR7 + (n - 7)


def main() -> int:
    if not SRC.exists():
        raise SystemExit(f"{SRC} not found")
    shutil.copyfile(SRC, DST)
    wb = openpyxl.load_workbook(DST)
    ws = wb["Sheet1"]

    def put(row: int, name: str, direction: str, note: str) -> None:
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=4).value = direction
        ws.cell(row=row, column=5).value = note

    # ---- the serial link itself: same four pads, wider frame ---------------
    put(ROW_MOSI, "MOSI", "Input",
        'bidir_PAD[1] (bond "config2"), north edge. Serial data in, MSB first: '
        "CMD[3:0] + ADDR[11:0] (+ DATA[7:0] or DATA[23:0] on writes). The command "
        "field widened from 2 to 4 bits for the 12-command set; the pad is unchanged.")
    put(ROW_MISO, "MISO", "Output",
        'bidir_PAD[3] (bond "config4"), north edge. Serial data out: RD_MET / RD_IMG / '
        "RD_ACT / RD_C all return 24 bits, MSB first.")

    # ---- status pads: same three, remapped -------------------------------
    put(ROW_BUSY, "busy", "Output",
        'bidir_PAD[4] (bond "config5"), east edge. Engine busy: high while the '
        "sequencer or the metrics unit is running. The host polls this between "
        "commands (writes and EXEC are only accepted while it is low).")
    put(ROW_DONE, "dla_busy", "Output",
        'bidir_PAD[5] (bond "config6"), east edge. MAC array active. Renamed from '
        '"done" on the main chip: the sequencer now owns the array, so array-level '
        "activity is the useful scope signal.")
    put(ROW_WBDONE, "dla_done", "Output",
        'bidir_PAD[6] (bond "config7"), east edge. MAC pass complete. Renamed from '
        '"wb_done"; same pad, same direction.')

    # ---- the ONE pin the discriminator costs ------------------------------
    put(bidir_row(7), "verdict", "Output",
        'bidir_PAD[7] (bond "config8"), east edge. THE ONLY PIN ADDED BY THE '
        "DISCRIMINATOR: high when D(generated) > 0.5, i.e. the generator fooled D. "
        "Convenience only - the same bit is readable over the link as MET_STATUS or "
        "derivable from MET_Y_FAKE, so this pad may be left unconnected with no loss "
        "of function. It exists so a scope can see the verdict without a serial read.")

    # ---- the eight pins the burst bus costs (nothing to do with D) --------
    for i in range(8):
        put(bidir_row(8 + i), f"pdata[{i}]", "Input",
            f'bidir_PAD[{8 + i}] (bond "config{9 + i}"), '
            f"{'east' if 8 + i < 12 else 'west'} edge. Parallel write bus for the "
            "WR_BURST8 command: one whole byte per SCLK edge instead of one bit. "
            "Throughput feature, unrelated to the discriminator - weight streaming is "
            "~99% of run time, and this takes a 1024-byte tile from 24,576 SCLK edges "
            "to 1,040 (23.6x). Tie low if only the serial commands are used.")

    # ---- readback redundancy ---------------------------------------------
    put(bidir_row(16), "MISO_mirror", "Output",
        'bidir_PAD[16] (bond "config17"), west edge. Second copy of MISO, driven from '
        "the same net by an independent pad driver. This chip has no scan chain, JTAG "
        "or BIST, so every readable value leaves through MISO; one open bond or damaged "
        "pad there makes the die unreadable. Bond this pad INSTEAD OF or AS WELL AS "
        "bidir_PAD[3] - shorting both at the board is safe.")

    # ---- summary lines ----------------------------------------------------
    ws.cell(row=ROW_SUMMARY, column=2).value = (
        "Summary: 91 bond pads total - 8 supply (single 3.3 V rail), 19 used digital "
        "signals (clk, rst_n, 4-wire serial link, 4 status outputs, 8-bit parallel "
        "write bus, 1 MISO mirror), 64 unused (3 bidir spares + 60 analog + 1 slot "
        "spare input). Versus the main chip this is +10 used pads, of which EIGHT are "
        "the parallel burst bus, ONE is readback redundancy and exactly ONE (verdict) "
        "is the discriminator - and the last two are optional. Minimum viable bring-up "
        "is 7 signals: clk, rst_n, SCLK, MOSI, CS_N, MISO, busy. The 60 analog pads cannot absorb digital signals: gf180mcu_fd_io__"
        "asig_5p0 exposes only a pass-through ASIG5V pin, with no A/Y/OE/IE.")
    ws.cell(row=ROW_PROTO, column=2).value = (
        "Serial protocol (12 commands, CMD[3:0]+ADDR[11:0] header): WR_A 0, WR_B 1, "
        "WR_IMG 2, WR_ACT 3, WR_CFG 4 (24-bit), EXEC 5 (ADDR={op,arg}), RD_MET 6, "
        "RD_IMG 7, RD_ACT 8, RD_C 9, WR_BURST 10 (auto-increment, 8 edges/byte), "
        "WR_BURST8 11 (parallel bus, 1 edge/byte). Full spec in EXPERIMENTAL_GAN_CHIP.md.")
    ws.cell(row=ROW_PLACE, column=2).value = (
        "Physical placement of every pad is fixed by the same wafer.space workshop-slot "
        "template as the main chip (die 2935 x 2935 um), so the bond map is unchanged - "
        "only the function assigned to four previously-spare bidir pads and the eight "
        "burst-bus pads differs. NOTE: unlike the main chip, this sheet does NOT reflect "
        "a signed-off GDS - place-and-route has not been run on gan_engine_top.")

    wb.save(DST)
    print(f"wrote {DST}")
    print("  main chip:         7 of 20 bidir pads used")
    print("  experimental chip: 17 of 20 bidir pads used")
    print("  attribution: +1 discriminator (verdict, optional), +8 parallel burst bus,")
    print("               +1 MISO mirror (readback redundancy, optional)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
